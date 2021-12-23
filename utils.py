import matplotlib.pyplot as plt
import numpy as np
import os
import openpyxl

ACC = [82.97, 53.46, 57.04, 92.84, 53.46, 78.52, 76.13, 74.94, 66.58, 85.68,
       57.04, 34.37, 88.06, 84.48, 97.99, 90.45, 79.71, 36.75, 68.97, 84.48,
       65.39, 48.68, 80.89, 95.22, 67.78, 91.64, 48.68, 41.52, 68.97, 49.88]

def calRecall(confusion):
    recall0 = confusion[0][0] / confusion[0].sum()
    recall1 = confusion[1][1] / confusion[1].sum()
    recall2 = confusion[2][2] / confusion[2].sum()
    return recall0, recall1, recall2

def calPrecision(confusion):
    prec0 = confusion[0][0] / (confusion[0][0] + confusion[1][0] + confusion[2][0])
    prec1 = confusion[1][1] / (confusion[0][1] + confusion[1][1] + confusion[2][1])
    prec2 = confusion[2][2] / (confusion[0][2] + confusion[1][2] + confusion[2][2])
    return prec0, prec1, prec2

def writeSubResults(file, accs, feature, size, colCnt):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['Sheet1']

    col = 2 + colCnt
    # 写入第一行，即feature + size
    title = feature + str(size)
    sheet.cell(1, col, value=title)
    # 写入每一个被试的数据
    for row in range(2, 32):
        sheet.cell(row, col, value=accs[row - 2])

    workbook.save(file)